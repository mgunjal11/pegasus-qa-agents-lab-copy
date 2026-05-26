#!/usr/bin/env python3
"""Generate QMetry-importable Excel matching the FF2.0 template layout."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Matches C:\Users\mgunjal\Downloads\QMetry FF2.0.xlsx
COLUMNS = [
    "Summary",
    "Automatable",
    "Automation Status",
    "Priority",
    "Folders",
    "Step Summary",
    "Test Type",
    "Status",
    "Regression Test (Y/N)",
    "Story",
    "TestData Dependent",
]

SHEET_NAME = "QMetry Template"

# Columns merged vertically for each 3-row test case (1-based). Step Summary (6) is never merged.
MERGE_COLUMNS = (1, 2, 3, 4, 5, 7, 8, 9, 10, 11)

# Column widths from sample template
COLUMN_WIDTHS = {
    "A": 50.63,
    "B": 14.82,
    "C": 15.82,
    "D": 9.54,
    "E": 16.18,
    "F": 89.0,
    "G": 11.73,
    "H": 8.73,
    "I": 22.27,
    "J": 16.0,
    "K": 22.36,
}


def read_tsv(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f, delimiter="\t"))


def normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        raise ValueError("TSV is empty")

    header = [c.strip() for c in rows[0]]
    if header == COLUMNS:
        return _fix_step_summary_rows(rows)

    # Legacy 20-column format -> map to QMetry 11-column rows
    legacy = {
        "Summary": "Summary",
        "Automatable": "Automatable",
        "Automation Status": "Automation Status",
        "Priority": "Priority",
        "Folders": "Folders",
        "Step Summary": "Step Summary",
        "Test Type": "Test Type",
        "Regression Test (Y/N)": "Regression Test (Y/N)",
        "Story": "Story",
        "TestData Dependent": "TestData Dependent",
    }
    if "Step Summary" in header:
        idx = {name: header.index(name) for name in legacy if name in header}
        out = [COLUMNS]
        for row in rows[1:]:
            if not any(cell.strip() for cell in row):
                continue
            mapped = []
            for col in COLUMNS:
                if col == "Status":
                    mapped.append("")
                elif col in idx and idx[col] < len(row):
                    mapped.append(row[idx[col]].strip())
                else:
                    mapped.append("")
            out.append(mapped)
        return _fix_step_summary_rows(out)

    raise ValueError(f"Unrecognized TSV header: {header[:5]}...")


def _fix_step_summary_rows(rows: list[list[str]]) -> list[list[str]]:
    """Ensure When:/Then: lines always land in the Step Summary column (index 5)."""
    step_idx = COLUMNS.index("Step Summary")
    fixed = [rows[0]]
    for row in rows[1:]:
        padded = row + [""] * (len(COLUMNS) - len(row))
        padded = padded[: len(COLUMNS)]
        step = padded[step_idx].strip()
        if not step:
            for cell in padded:
                text = cell.strip()
                if text.startswith(("When:", "Then:", "Given:")):
                    step = text
                    break
        if step.startswith(("When:", "Then:")) and not padded[0].strip():
            padded = [""] * len(COLUMNS)
            padded[step_idx] = step
        elif step.startswith("Given:") and padded[0].strip():
            padded[step_idx] = step
        else:
            padded[step_idx] = step
        fixed.append(padded)
    return fixed


def write_xlsx(rows: list[list[str]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    data_rows = rows[1:]
    for row_idx, row in enumerate(data_rows, start=2):
        padded = row + [""] * (len(COLUMNS) - len(row))
        for col_idx, value in enumerate(padded[: len(COLUMNS)], start=1):
            cell_value = value if value else None
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # Merge metadata cells for each 3-row Given/When/Then block
    if len(data_rows) >= 3:
        start = 2
        while start <= len(data_rows) + 1:
            end = min(start + 2, len(data_rows) + 1)
            if end - start + 1 == 3:
                for col in MERGE_COLUMNS:
                    ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
            start += 3

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python generate_qmetry_excel.py <input.tsv> [output.xlsx]", file=sys.stderr)
        return 1

    tsv_path = Path(sys.argv[1])
    if not tsv_path.exists():
        print(f"Error: file not found: {tsv_path}", file=sys.stderr)
        return 1

    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else tsv_path.with_suffix(".xlsx")
    rows = normalize_rows(read_tsv(tsv_path))
    write_xlsx(rows, xlsx_path)
    print(str(xlsx_path.resolve()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
