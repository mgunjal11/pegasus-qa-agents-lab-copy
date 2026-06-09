#!/usr/bin/env python3
"""
Fetch and parse test plans for msc-dev-code-and-qa-test-coverage-validator.

Sources (priority order):
  1. --attachment local file(s)
  2. Jira issue attachments (REST API)
  3. SharePoint / comment-referenced Excel in testplans/ (see jira cache testPlanReferences)
  4. testcases/{KEY}-testcases.xlsx|.tsv
  5. .coverage-validator.defaults.json / manifest testPlanPath + testPlanSheet

Supports QMetry sheets and Domino-style evidence sheets (e.g. "Inc as full").

  python scripts/fetch_jira_testplan.py MSC-205625 --from-jira-cache
  python scripts/fetch_jira_testplan.py MSC-205625 --sheet "Inc as full"
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from jira_env import (
    credentials_hint,
    download_attachment,
    fetch_issue_attachments,
    load_dotenv,
)
from confluence_requirements import (
    build_ladr_traceability,
    compute_testplan_coverage,
    dedupe_ladr_requirements,
    fetch_and_cache_confluence_for_issue,
    format_testplan_coverage_detail,
    load_confluence_cache,
    map_testcases_to_requirements,
    merge_requirement_sets,
)
from testplan_evidence import extract_testcase_evidence_ids, row_evidence_text
from testplan_gwt import (
    gwt_key_from_marker,
    has_complete_gwt,
    is_step_blob_column,
    merge_steps,
    normalize_gwt_typos,
    normalize_steps,
    parse_gwt_from_text,
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[misc, assignment]

QMETRY_COLUMNS = [
    "Summary",
    "Automatable",
    "Automation Status",
    "Priority",
    "Folders",
    "Step Summary",
    "Test Type",
    "Status",
    "Regression Test (Y/N)",
    "Story",
    "TestData Dependent",
]

TESTPLAN_NAME_RE = re.compile(
    r"(test[\s_-]?case|test[\s_-]?plan|qmetry|testplan|testcases|domino)",
    re.IGNORECASE,
)
TESTPLAN_EXT = {".xlsx", ".xls", ".tsv", ".csv", ".txt"}
REQ_ID_RE = re.compile(r"\bR(\d+)\b", re.IGNORECASE)
GWT_RE = re.compile(r"^(Given|When|Then|Than|Tehn|Them)\s*:?\s*", re.IGNORECASE)
ISSUE_KEY_RE = re.compile(r"\b(MSC-\d+)\b", re.IGNORECASE)
SHAREPOINT_URL_RE = re.compile(r"https://[^\s\"']*sharepoint\.com[^\s\"']*", re.IGNORECASE)
XLSX_FILE_PARAM_RE = re.compile(r"file=([^&\"'\s]+\.xlsx)", re.IGNORECASE)
SHEET_REF_RE = re.compile(
    r"(?:Refer(?:ence)?|see|use|check)\s+(?:the\s+)?"
    r"(?:[\"']?([^\"']+?)[\"']?\s+sheet|sheet\s+[\"']?([^\"']+?)[\"']?)",
    re.IGNORECASE,
)
INC_AS_FULL_RE = re.compile(r"\bInc\s+as\s+full\b", re.IGNORECASE)
TEST_PLAN_LINK_TEXT_RE = re.compile(
    r"test\s*plan\s*(?:and\s*evidence)?|testplan",
    re.IGNORECASE,
)

MASCOT_URL_RE = re.compile(
    r"https://[^\s\"']*foundry\.wbdapps\.com/mascot[^\s\"']*",
    re.IGNORECASE,
)

EVIDENCE_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "summary": (
        "summary",
        "test scenario",
        "scenario",
        "test case",
        "description",
        "test name",
        "objective",
        "test objective",
    ),
    "given": ("preconditions", "precondition", "given", "prerequisites", "setup"),
    "when": ("when",),
    "then": ("expected", "expected result", "then", "verification", "expected outcome"),
    "story": ("story", "jira", "jira id", "ticket", "issue", "issue key", "jira key", "bug"),
    "status": ("status", "result", "pass/fail", "execution status", "test status"),
    "priority": ("priority",),
    "test_type": ("test type", "type"),
}


@dataclass
class TestCase:
    id: str
    summary: str
    story: str
    priority: str
    test_type: str
    automatable: str
    regression: str
    steps: dict[str, str] = field(default_factory=dict)
    mapped_requirements: list[str] = field(default_factory=list)
    source_file: str = ""
    source_sheet: str = ""
    section: str = ""
    comment: str = ""
    evidence_text: str = ""
    mascot_links: list[dict[str, str]] = field(default_factory=list)
    evidence_ids: list[dict[str, str]] = field(default_factory=list)


def _xlsx_cell_text(cell: Any) -> str:
    """Cell value; use Excel hyperlink target when display text is e.g. 'Mascot'."""
    val = "" if cell.value is None else str(cell.value).strip()
    target = ""
    hl = getattr(cell, "hyperlink", None)
    if hl is not None and getattr(hl, "target", None):
        target = str(hl.target).strip()
    if target and MASCOT_URL_RE.search(target):
        return target
    if val and MASCOT_URL_RE.search(val):
        return val
    if target and val.lower() in {"mascot", "link", "url", "open"}:
        return target
    return val


def _mascot_column_label(header_name: str) -> str:
    return re.sub(r"\s+", " ", header_name.replace("QA-", "").replace("SIT-", "").strip())


def extract_mascot_links(header: list[str], row: list[str]) -> list[dict[str, str]]:
    """Pull Mascot fulfillment URLs from QA/SIT mascot columns and any row cell."""
    links: list[dict[str, str]] = []
    seen: set[str] = set()

    def add_urls(cell: str, label: str) -> None:
        if not cell:
            return
        for url in MASCOT_URL_RE.findall(cell):
            if url not in seen:
                seen.add(url)
                links.append({"label": label or "Mascot", "url": url})

    for idx, name in enumerate(header):
        if "mascot" not in name.lower():
            continue
        cell = row[idx].strip() if idx < len(row) else ""
        add_urls(cell, _mascot_column_label(name))

    # Fallback: URLs in other columns (e.g. Comments, Evidence) not named mascot
    for idx, cell in enumerate(row):
        if idx < len(header) and "mascot" in header[idx].lower():
            continue
        add_urls(cell.strip() if cell else "", "Evidence")

    return links


def scenario_label(tc: TestCase) -> str:
    """High-level scenario label: Section · Summary when section is present."""
    if tc.section and tc.summary:
        return f"{tc.section} · {tc.summary}"
    return tc.section or tc.summary or tc.id


def format_testplan_summary_note(
    issue_key: str,
    filename: str,
    comment_sheet: str | None,
    excel_sheet: str | None,
    scenario_count: int,
    scenario_filter: str | None = None,
) -> str:
    """Standard note-box text for section 3 when test plan is parsed from Jira attachment."""
    sheet_from_comment = comment_sheet or "Inc as full"
    excel_tab = excel_sheet or sheet_from_comment
    filter_note = scenario_filter.strip() if scenario_filter else ""
    count_part = f"{scenario_count} {filter_note}".strip() if filter_note else str(scenario_count)
    return (
        f"Downloaded {filename} from Jira attachment comment sheet {sheet_from_comment} "
        f"→ Excel tab {excel_tab} · {count_part} scenarios for {issue_key}."
    )

def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def cache_path(issue_key: str) -> Path:
    return repo_root() / "reports" / ".cache" / f"{issue_key.upper()}-testplan.json"


def jira_cache_path(issue_key: str) -> Path:
    return repo_root() / "reports" / ".cache" / f"{issue_key.upper()}-jira.json"


def manifest_path(issue_key: str) -> Path:
    return repo_root() / "reports" / ".cache" / f"{issue_key.upper()}-manifest.json"


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_run_options(issue_key: str) -> dict[str, Any]:
    opts: dict[str, Any] = {}
    defaults = load_json(repo_root() / ".coverage-validator.defaults.json")
    manifest = load_json(manifest_path(issue_key))
    opts.update(defaults)
    opts.update(manifest)
    return opts


def adf_to_text(node: Any) -> str:
    """Flatten Atlassian Document Format to plain text."""
    if node is None:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return " ".join(adf_to_text(n) for n in node)
    if not isinstance(node, dict):
        return str(node)
    parts: list[str] = []
    if node.get("type") == "text" and node.get("text"):
        parts.append(str(node["text"]))
    for key in ("content", "marks"):
        if key in node:
            parts.append(adf_to_text(node[key]))
    return " ".join(p for p in parts if p)


def collect_comment_texts(jira_data: dict[str, Any]) -> list[str]:
    texts: list[str] = []
    for key in ("comments", "commentTexts"):
        raw = jira_data.get(key)
        if isinstance(raw, list):
            for item in raw:
                if isinstance(item, str):
                    texts.append(item)
                elif isinstance(item, dict):
                    for field in ("body", "text", "markdown"):
                        if field in item and isinstance(item[field], str):
                            texts.append(item[field])
                        elif field == "body" and isinstance(item.get("body"), dict):
                            texts.append(adf_to_text(item["body"]))
    fields = jira_data.get("fields") or {}
    comment_block = fields.get("comment") or {}
    if isinstance(comment_block, dict):
        for item in comment_block.get("comments") or []:
            if isinstance(item, dict):
                body = item.get("body")
                if isinstance(body, str):
                    texts.append(body)
                elif isinstance(body, dict):
                    texts.append(adf_to_text(body))
    if isinstance(jira_data.get("description"), str):
        texts.append(jira_data["description"])
    fields_desc = fields.get("description")
    if isinstance(fields_desc, str):
        texts.append(fields_desc)
    return texts


def extract_sheet_name(text: str) -> str | None:
    if INC_AS_FULL_RE.search(text):
        return "Inc as full"
    match = SHEET_REF_RE.search(text)
    if match:
        name = (match.group(1) or match.group(2) or "").strip()
        if name:
            return name
    return None


def extract_sharepoint_filename(url: str) -> str | None:
    match = XLSX_FILE_PARAM_RE.search(url)
    if match:
        from urllib.parse import unquote

        return unquote(match.group(1)).replace("+", " ")
    if url.lower().endswith(".xlsx"):
        return Path(url.split("?")[0]).name
    return None


def extract_testplan_references(jira_data: dict[str, Any], issue_key: str) -> list[dict[str, Any]]:
    """Find test plan Excel + sheet references from Jira cache comments and description."""
    refs: list[dict[str, Any]] = []
    seen: set[str] = set()

    def add_ref(ref: dict[str, Any]) -> None:
        key = f"{ref.get('url','')}|{ref.get('filename','')}|{ref.get('sheet','')}"
        if key not in seen:
            seen.add(key)
            refs.append(ref)

    for existing in jira_data.get("testPlanReferences") or []:
        if isinstance(existing, dict):
            add_ref(existing)
    single = jira_data.get("testPlanReference")
    if isinstance(single, dict):
        enriched = dict(single)
        if not enriched.get("filename"):
            enriched["filename"] = extract_sharepoint_filename(enriched.get("url", "")) or "Domino Test Plan.xlsx"
        for text in collect_comment_texts(jira_data):
            sheet = extract_sheet_name(text)
            if sheet:
                enriched.setdefault("sheet", sheet)
                break
        add_ref(enriched)

    for text in collect_comment_texts(jira_data):
        sheet = extract_sheet_name(text)
        urls = SHAREPOINT_URL_RE.findall(text)
        if not urls and TEST_PLAN_LINK_TEXT_RE.search(text):
            urls = re.findall(r"https://[^\s\"']+", text)
        for url in urls:
            filename = extract_sharepoint_filename(url) or "Domino Test Plan.xlsx"
            add_ref(
                {
                    "type": "sharepoint",
                    "source": "jira_comment",
                    "url": url,
                    "filename": filename,
                    "sheet": sheet or jira_data.get("testPlanSheet"),
                    "linkText": "Test plan and evidence" if TEST_PLAN_LINK_TEXT_RE.search(text) else None,
                    "issueKey": issue_key,
                }
            )
        if sheet and not urls:
            add_ref(
                {
                    "type": "sheet_reference",
                    "source": "jira_comment",
                    "filename": "Domino Test Plan.xlsx",
                    "sheet": sheet,
                    "issueKey": issue_key,
                }
            )

    return refs


def is_testplan_attachment(filename: str) -> bool:
    name = filename.lower()
    ext = Path(name).suffix
    if ext not in TESTPLAN_EXT:
        return False
    if TESTPLAN_NAME_RE.search(name):
        return True
    if ext in {".xlsx", ".tsv"} and re.search(r"MSC-\d+", name, re.IGNORECASE):
        return True
    return ext in {".xlsx", ".tsv"} and ("test" in name or "case" in name)


def attachments_from_jira_cache(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    direct = data.get("attachments")
    if isinstance(direct, list):
        return direct
    fields = data.get("fields") or {}
    if isinstance(fields, dict):
        att = fields.get("attachment")
        if isinstance(att, list):
            return att
    return []


def resolve_local_file(
    issue_key: str,
    ref: dict[str, Any],
    run_opts: dict[str, Any],
) -> tuple[Path | None, str | None]:
    """Resolve referenced Excel to a local path and sheet name."""
    sheet = (
        run_opts.get("testPlanSheet")
        or ref.get("sheet")
        or "Inc as full"
        if ref.get("type") in {"sharepoint", "sheet_reference"}
        else ref.get("sheet")
    )
    filename = (
        run_opts.get("testPlanFilename")
        or ref.get("filename")
        or "Domino Test Plan.xlsx"
    )
    explicit = run_opts.get("testPlanPath")
    if explicit:
        path = Path(explicit)
        if not path.is_absolute():
            path = repo_root() / path
        if path.exists():
            return path, sheet or run_opts.get("testPlanSheet")

    root = repo_root()
    candidates = [
        root / "testplans" / filename,
        root / "testplans" / issue_key / filename,
        root / "testplans" / f"{issue_key}-testplan.xlsx",
        root / "testplans" / f"{issue_key}-testplan.xls",
        root / "testcases" / f"{issue_key}-testcases.xlsx",
        root / "testcases" / f"{issue_key}-testcases.tsv",
        root / "testcases" / filename,
        root / "testcases" / "Domino Test Plan.xlsx",
        root / "reports" / ".cache" / f"{issue_key}-testplan-files" / filename,
    ]
    for path in candidates:
        if path.exists():
            return path, sheet
    return None, sheet


def read_tsv_rows(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f, delimiter="\t"))


def normalize_qmetry_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    header = [c.strip() for c in rows[0]]
    if header == QMETRY_COLUMNS:
        return _fix_step_summary_rows(rows)
    col_idx = {name: header.index(name) for name in QMETRY_COLUMNS if name in header}
    if "Summary" not in col_idx or "Step Summary" not in col_idx:
        return rows
    out = [QMETRY_COLUMNS]
    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue
        padded = row + [""] * (len(header) - len(row))
        mapped = []
        for col in QMETRY_COLUMNS:
            idx = col_idx.get(col)
            if col == "Status":
                mapped.append("")
            elif idx is not None and idx < len(padded):
                mapped.append(padded[idx].strip())
            else:
                mapped.append("")
        out.append(mapped)
    return _fix_step_summary_rows(out)


def _fix_step_summary_rows(rows: list[list[str]]) -> list[list[str]]:
    step_idx = QMETRY_COLUMNS.index("Step Summary")
    fixed = [rows[0] if rows[0] == QMETRY_COLUMNS else QMETRY_COLUMNS]
    for row in rows[1:]:
        padded = row + [""] * (len(QMETRY_COLUMNS) - len(row))
        padded = padded[: len(QMETRY_COLUMNS)]
        step = padded[step_idx].strip()
        if not step:
            for cell in padded:
                text = cell.strip()
                if GWT_RE.match(text):
                    step = text
                    break
        step_match = GWT_RE.match(step)
        if step_match and step_match.group(1).lower() in ("when", "then"):
            padded[0] = ""
            padded[step_idx] = step
        elif step_match and step_match.group(1).lower() == "given" and padded[0].strip():
            padded[step_idx] = step
        else:
            padded[step_idx] = step
        fixed.append(padded)
    return fixed


def _cell(row: list[str], col_idx: dict[str, int], name: str) -> str:
    idx = col_idx.get(name)
    if idx is None or idx >= len(row):
        return ""
    return row[idx].strip()


def _assign_step(tc: TestCase, text: str) -> None:
    if not text:
        return
    text = normalize_gwt_typos(text)
    parsed = parse_gwt_from_text(text)
    if len(parsed) >= 2:
        tc.steps = merge_steps(tc.steps, parsed)
        return
    match = GWT_RE.match(text.strip())
    if match:
        key = gwt_key_from_marker(match.group(1))
        tc.steps[key] = text.strip()
    elif "given" not in tc.steps:
        tc.steps["given"] = text.strip()


def _ingest_step_columns(
    tc: TestCase,
    header: list[str],
    padded: list[str],
    cols: dict[str, int],
) -> None:
    """Load steps from dedicated GWT columns and any step-blob column (content-based)."""

    def col(name: str) -> str:
        idx = cols.get(name)
        if idx is None or idx >= len(padded):
            return ""
        return padded[idx].strip()

    blobs: list[str] = []
    for idx, name in enumerate(header):
        if is_step_blob_column(name):
            cell = padded[idx].strip() if idx < len(padded) else ""
            if cell:
                blobs.append(cell)
    for key in ("given", "when", "then"):
        val = col(key)
        if val:
            blobs.append(val)
    for blob in blobs:
        parsed = parse_gwt_from_text(blob)
        if parsed:
            tc.steps = merge_steps(tc.steps, parsed)
        else:
            _assign_step(tc, blob)
    tc.steps = normalize_steps(tc.steps)


def _summary_from_steps(tc: TestCase) -> str:
    given = tc.steps.get("given", "")
    if given.lower().startswith("given:"):
        return given[6:].strip()[:120]
    return given[:120] if given else f"Test case {tc.id}"


def rows_to_cases(rows: list[list[str]], source: str, sheet: str = "") -> list[TestCase]:
    if not rows:
        return []
    rows = normalize_qmetry_rows(rows)
    header = [c.strip() for c in rows[0]]
    col_idx = {name: header.index(name) for name in QMETRY_COLUMNS if name in header}
    if "Summary" not in col_idx or "Step Summary" not in col_idx:
        raise ValueError(f"Unrecognized QMetry header in {source}: {header[:6]}")

    cases: list[TestCase] = []
    current: TestCase | None = None
    tc_num = 0

    def flush() -> None:
        nonlocal current
        if current and (current.summary or any(current.steps.values())):
            cases.append(current)
        current = None

    for row in rows[1:]:
        padded = row + [""] * (len(QMETRY_COLUMNS) - len(row))
        padded = padded[: len(QMETRY_COLUMNS)]
        summary = padded[col_idx["Summary"]].strip()
        step_raw = padded[col_idx["Step Summary"]].strip()
        if GWT_RE.match(summary) and not GWT_RE.match(step_raw):
            step_raw = summary
            summary = ""
        if summary:
            flush()
            tc_num += 1
            current = TestCase(
                id=f"TC{tc_num}",
                summary=summary,
                story=_cell(padded, col_idx, "Story"),
                priority=_cell(padded, col_idx, "Priority"),
                test_type=_cell(padded, col_idx, "Test Type"),
                automatable=_cell(padded, col_idx, "Automatable"),
                regression=_cell(padded, col_idx, "Regression Test (Y/N)"),
                source_file=source,
                source_sheet=sheet,
            )
            if step_raw:
                _assign_step(current, step_raw)
        elif current is not None:
            _assign_step(current, step_raw or summary)
        elif step_raw:
            tc_num += 1
            current = TestCase(
                id=f"TC{tc_num}",
                summary="",
                story="",
                priority="",
                test_type="",
                automatable="",
                regression="",
                source_file=source,
                source_sheet=sheet,
            )
            _assign_step(current, step_raw)

    flush()
    for tc in cases:
        tc.steps = normalize_steps(tc.steps)
        if not tc.summary:
            tc.summary = _summary_from_steps(tc)
    return [tc for tc in cases if tc.summary or tc.steps]


def _find_evidence_columns(header: list[str]) -> dict[str, int]:
    lower = [h.lower().strip() for h in header]
    found: dict[str, int] = {}
    for key, aliases in EVIDENCE_COLUMN_ALIASES.items():
        for idx, name in enumerate(lower):
            if name in aliases or any(a in name for a in aliases):
                found[key] = idx
                break
    return found


def parse_evidence_rows(
    rows: list[list[str]],
    source: str,
    sheet: str,
    issue_key: str | None = None,
) -> list[TestCase]:
    """Parse Domino-style evidence sheets (e.g. Inc as full)."""
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    cols = _find_evidence_columns(header)
    if "summary" not in cols and len(header) > 0:
        cols["summary"] = 0

    cases: list[TestCase] = []
    tc_num = 0
    for row in rows[1:]:
        if not any(str(c or "").strip() for c in row):
            continue
        padded = [str(c or "").strip() for c in row] + [""] * (len(header) - len(row))

        def col(name: str) -> str:
            idx = cols.get(name)
            if idx is None or idx >= len(padded):
                return ""
            return padded[idx].strip()

        summary = col("summary")
        story = col("story")
        row_text = " ".join(padded).lower()

        if issue_key:
            key_upper = issue_key.upper()
            if story and key_upper not in story.upper() and key_upper not in row_text.upper():
                if ISSUE_KEY_RE.search(row_text):
                    keys_in_row = ISSUE_KEY_RE.findall(row_text.upper())
                    if key_upper not in keys_in_row:
                        continue
                elif story:
                    continue

        step_blob = " ".join(
            padded[idx].strip()
            for idx, name in enumerate(header)
            if is_step_blob_column(name) and idx < len(padded) and padded[idx].strip()
        )
        if not summary and not step_blob and not any(col(k) for k in ("given", "when", "then")):
            continue

        tc_num += 1
        tc = TestCase(
            id=f"TC{tc_num}",
            summary=summary or f"Row {tc_num + 1}",
            story=story or issue_key or "",
            priority=col("priority"),
            test_type=col("test_type") or "End to End",
            automatable="",
            regression="",
            source_file=source,
            source_sheet=sheet,
            comment=col("comment") if "comment" in cols else "",
            evidence_text=row_evidence_text(header, padded),
            mascot_links=extract_mascot_links(header, padded),
        )
        _ingest_step_columns(tc, header, padded, cols)
        if not tc.summary or tc.summary.startswith("Row "):
            tc.summary = _summary_from_steps(tc) if tc.steps else (summary or f"Row {tc_num + 1}")
        cases.append(tc)
    return cases


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def _sheet_name_aliases(name: str | None) -> list[str]:
    if not name:
        return []
    base = _normalize_sheet_name(name)
    aliases = [base]
    if "inc as full" in base:
        aliases.extend(["inc as fulll", "inc as full"])
    return aliases


def _pick_worksheet(wb: Any, sheet_name: str | None) -> Any:
    if sheet_name:
        for alias in _sheet_name_aliases(sheet_name):
            for ws_name in wb.sheetnames:
                normalized = _normalize_sheet_name(ws_name)
                if normalized == alias or alias in normalized:
                    return wb[ws_name]
    if "QMetry Template" in wb.sheetnames:
        return wb["QMetry Template"]
    return wb.active


def read_xlsx_sheet_rows(path: Path, sheet_name: str | None) -> tuple[list[list[str]], str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Run: pip install openpyxl")
    # read_only=False so Excel hyperlinks (display text "Mascot") resolve to URLs
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = _pick_worksheet(wb, sheet_name)
    actual_sheet = ws.title
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        rows.append([_xlsx_cell_text(cell) for cell in row])
    wb.close()
    return rows, actual_sheet


def _header_has_serial_column(lower: list[str]) -> bool:
    return any(re.sub(r"[\s._-]", "", h) in {"srno", "srnumber", "sno"} for h in lower)


def is_step_summary_sheet(header: list[str]) -> bool:
    """QMetry / Domino layout: Summary + Step Summary (column names may vary)."""
    lower = [h.lower().strip() for h in header]
    return "summary" in lower and any("step summary" in h for h in lower)


def is_domino_testplan_header(header: list[str]) -> bool:
    lower = [h.lower().strip() for h in header]
    return is_step_summary_sheet(header) and _header_has_serial_column(lower)


def parse_domino_rows(
    rows: list[list[str]],
    source: str,
    sheet: str,
    issue_key: str | None = None,
) -> list[TestCase]:
    """Parse Domino Test Plan sheets (Summary + Step Summary columns, 3-row GWT blocks)."""
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    lower = [h.lower().strip() for h in header]
    summary_idx = lower.index("summary") if "summary" in lower else 2
    step_idx = next(i for i, h in enumerate(lower) if "step summary" in h)
    section_idx = lower.index("section") if "section" in lower else None
    comment_idx = lower.index("comment") if "comment" in lower else None
    story_idx = next((i for i, h in enumerate(lower) if h.strip() == "story"), None)
    priority_idx = next((i for i, h in enumerate(lower) if "priority" in h), None)
    type_idx = next((i for i, h in enumerate(lower) if "test type" in h), None)

    cases: list[TestCase] = []
    current: TestCase | None = None
    tc_num = 0

    def flush() -> None:
        nonlocal current
        if current and (current.summary or any(current.steps.values())):
            cases.append(current)
        current = None

    def cell(row: list[str], idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx] or "").strip()

    for row in rows[1:]:
        padded = [str(c or "").strip() for c in row] + [""] * (len(header) - len(row))
        summary = cell(padded, summary_idx)
        step_raw = cell(padded, step_idx)
        if GWT_RE.match(summary) and not GWT_RE.match(step_raw):
            step_raw = summary
            summary = ""
        if summary:
            flush()
            tc_num += 1
            current = TestCase(
                id=f"TC{tc_num}",
                summary=summary,
                story=cell(padded, story_idx) or issue_key or "",
                priority=cell(padded, priority_idx),
                test_type=cell(padded, type_idx) or "End to End",
                automatable=cell(padded, lower.index("automatable")) if "automatable" in lower else "",
                regression=cell(padded, next((i for i, h in enumerate(lower) if "regression" in h), None)) or "",
                source_file=source,
                source_sheet=sheet,
                section=cell(padded, section_idx),
                comment=cell(padded, comment_idx),
                evidence_text=row_evidence_text(header, padded),
                mascot_links=extract_mascot_links(header, padded),
            )
            if step_raw:
                _assign_step(current, step_raw)
        elif current is not None:
            _assign_step(current, step_raw or summary)
        elif step_raw:
            tc_num += 1
            current = TestCase(
                id=f"TC{tc_num}",
                summary="",
                story=issue_key or "",
                priority="",
                test_type="",
                automatable="",
                regression="",
                source_file=source,
                source_sheet=sheet,
            )
            _assign_step(current, step_raw)

    flush()
    for tc in cases:
        tc.steps = normalize_steps(tc.steps)
        if not tc.summary:
            tc.summary = _summary_from_steps(tc)

    if issue_key:
        req_blob = issue_key.lower()
        filtered = []
        for tc in cases:
            hay = " ".join([tc.summary, " ".join(tc.steps.values())]).lower()
            if issue_key.upper() in hay.upper():
                filtered.append(tc)
                continue
            if "passport" in req_blob or issue_key.upper().startswith("MSC-"):
                if "passport" in hay and (
                    "incremental as full" in hay
                    or "incremental as full" in tc.summary.lower()
                    or "full ff" in hay
                    or "full fulfillment" in hay
                ):
                    filtered.append(tc)
        if filtered:
            for i, tc in enumerate(filtered, 1):
                tc.id = f"TC{i}"
            return filtered
    return cases


def is_qmetry_header(header: list[str]) -> bool:
    stripped = [c.strip() for c in header[: len(QMETRY_COLUMNS)]]
    return stripped == QMETRY_COLUMNS


def read_xlsx_cases(path: Path, sheet_name: str | None = None, issue_key: str | None = None) -> list[TestCase]:
    rows, actual_sheet = read_xlsx_sheet_rows(path, sheet_name)
    if not rows:
        return []
    header = [c.strip() for c in rows[0]]
    if is_domino_testplan_header(header) or is_step_summary_sheet(header):
        return parse_domino_rows(rows, path.name, actual_sheet, issue_key)
    if is_qmetry_header(header):
        if is_qmetry_header(header):
            normalized = [QMETRY_COLUMNS]
            for row in rows[1:]:
                padded = row + [""] * (len(QMETRY_COLUMNS) - len(row))
                normalized.append(padded[: len(QMETRY_COLUMNS)])
            return rows_to_cases(normalized, path.name, actual_sheet)
        col_idx = {name: header.index(name) for name in QMETRY_COLUMNS if name in header}
        if "Summary" in col_idx and "Step Summary" in col_idx:
            return rows_to_cases(normalize_qmetry_rows(rows), path.name, actual_sheet)
    return parse_evidence_rows(rows, path.name, actual_sheet, issue_key)


def parse_testplan_file(path: Path, sheet_name: str | None = None, issue_key: str | None = None) -> list[TestCase]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return read_xlsx_cases(path, sheet_name, issue_key)
    if ext in {".tsv", ".txt"}:
        return rows_to_cases(read_tsv_rows(path), path.name)
    if ext == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        return rows_to_cases(rows, path.name)
    raise ValueError(f"Unsupported test plan format: {path}")


def extract_requirements(jira_cache: Path | None) -> list[dict[str, str]]:
    if not jira_cache or not jira_cache.exists():
        return []
    data = json.loads(jira_cache.read_text(encoding="utf-8"))
    reqs = data.get("requirements") or []
    out: list[dict[str, str]] = []
    for item in reqs:
        if isinstance(item, str):
            match = re.match(r"^(R\d+)\s*:\s*(.+)$", item.strip(), re.IGNORECASE)
            if match:
                out.append({"id": match.group(1).upper(), "text": match.group(2).strip()})
            else:
                out.append({"id": f"R{len(out)+1}", "text": item.strip()})
        elif isinstance(item, dict):
            rid = str(item.get("id") or f"R{len(out)+1}").upper()
            out.append({"id": rid, "text": str(item.get("text") or "").strip()})
    return out


def resolve_testplan_files(
    issue_key: str,
    from_jira_cache: bool,
    local_paths: list[str],
    sheet_override: str | None,
    site: str,
) -> tuple[list[tuple[Path, str | None]], list[dict[str, Any]], list[dict[str, Any]], str | None]:
    dest_dir = repo_root() / "reports" / ".cache" / f"{issue_key.upper()}-testplan-files"
    downloaded: list[tuple[Path, str | None]] = []
    meta: list[dict[str, Any]] = []
    auth_error: str | None = None
    jira_cache = jira_cache_path(issue_key)
    jira_data = load_json(jira_cache) if from_jira_cache or jira_cache.exists() else {}
    run_opts = load_run_options(issue_key)
    testplan_refs = extract_testplan_references(jira_data, issue_key)

    if local_paths:
        sheet = sheet_override or run_opts.get("testPlanSheet")
        for p in local_paths:
            path = Path(p)
            if path.exists():
                downloaded.append((path, sheet))
                meta.append({"filename": path.name, "local": True, "sheet": sheet})
        return downloaded, meta, testplan_refs, None

    attachments: list[dict[str, Any]] = []
    if from_jira_cache or jira_cache.exists():
        attachments = attachments_from_jira_cache(jira_cache)
    attachments = [a for a in attachments if a.get("content")]

    if not attachments:
        try:
            attachments = fetch_issue_attachments(issue_key, site)
        except RuntimeError as exc:
            auth_error = str(exc)
            attachments = [a for a in attachments_from_jira_cache(jira_cache) if a.get("content")]

    ref_sheet = (
        sheet_override
        or run_opts.get("testPlanSheet")
        or (testplan_refs[0].get("sheet") if testplan_refs else None)
    )
    candidates = [a for a in attachments if is_testplan_attachment(a.get("filename", ""))]
    for att in candidates:
        try:
            path = download_attachment(att, dest_dir)
            downloaded.append((path, ref_sheet))
            meta.append(
                {
                    "id": att.get("id"),
                    "filename": att.get("filename"),
                    "source": "jira_attachment",
                    "sheet": ref_sheet,
                    "localPath": str(path),
                    "localFound": True,
                }
            )
        except RuntimeError as exc:
            auth_error = str(exc)

    if not downloaded and testplan_refs:
        att_names = {a.get("filename") for a in attachments if a.get("filename")}
        refs_to_use = testplan_refs
        if att_names:
            matched = [r for r in testplan_refs if r.get("filename") in att_names]
            if matched:
                refs_to_use = matched
        seen_paths: set[str] = set()
        for ref in refs_to_use:
            path, sheet = resolve_local_file(issue_key, ref, run_opts)
            sheet = sheet_override or sheet
            meta.append(
                {
                    "source": ref.get("source", "reference"),
                    "type": ref.get("type"),
                    "filename": ref.get("filename"),
                    "sheet": sheet,
                    "url": ref.get("url"),
                    "localFound": bool(path and path.exists()),
                    "localPath": str(path) if path else None,
                }
            )
            if path and path.exists():
                key = str(path.resolve())
                if key not in seen_paths:
                    seen_paths.add(key)
                    downloaded.append((path, sheet))

    if not downloaded:
        root = repo_root()
        fallbacks = [
            (root / "testcases" / f"{issue_key}-testcases.xlsx", sheet_override),
            (root / "testcases" / f"{issue_key}-testcases.tsv", None),
        ]
        for path, sheet in fallbacks:
            if path.exists():
                downloaded.append((path, sheet))
                meta.append({"filename": path.name, "local": True, "source": "workspace_fallback"})

    return downloaded, meta, testplan_refs, auth_error


def enrich_jira_cache_testplan_refs(issue_key: str) -> list[dict[str, Any]]:
    """Update jira cache with extracted testPlanReferences from stored comments."""
    path = jira_cache_path(issue_key)
    if not path.exists():
        return []
    data = load_json(path)
    refs = extract_testplan_references(data, issue_key)
    # Prefer comment-sourced refs with filename + sheet over bare URL-only entries
    refs.sort(
        key=lambda r: (
            0 if r.get("filename") and r.get("sheet") else 1,
            0 if r.get("source") == "jira_comment" else 1,
        )
    )
    if refs:
        data["testPlanReferences"] = refs
        data["testPlanReference"] = refs[0]
        path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    return refs


def main() -> int:
    parser = argparse.ArgumentParser(description="Fetch and parse Jira / referenced test plans")
    parser.add_argument("issue_key", help="Jira issue key, e.g. MSC-205625")
    parser.add_argument("--from-jira-cache", action="store_true")
    parser.add_argument("--attachment", action="append", default=[], help="Local test plan file")
    parser.add_argument("--sheet", help='Excel sheet name, e.g. "Inc as full"')
    parser.add_argument("--site", default="wbdstreaming.atlassian.net")
    parser.add_argument("--output", help="Output JSON path")
    args = parser.parse_args()

    load_dotenv()
    issue_key = args.issue_key.upper()
    out_path = Path(args.output) if args.output else cache_path(issue_key)
    jira_cache = jira_cache_path(issue_key)

    enrich_jira_cache_testplan_refs(issue_key)

    files, attachment_meta, testplan_refs, auth_error = resolve_testplan_files(
        issue_key, args.from_jira_cache, args.attachment, args.sheet, args.site
    )

    jira_files = [
        m.get("filename")
        for m in attachment_meta
        if m.get("source") == "jira_attachment" and m.get("localFound", True)
    ]
    if jira_files and not args.attachment:
        allowed = set(jira_files)
        files = [(p, s) for p, s in files if p.name in allowed]

    all_cases: list[TestCase] = []
    parse_errors: list[str] = []
    sheets_used: list[str] = []
    for path, sheet in files:
        try:
            cases = parse_testplan_file(path, sheet, issue_key)
            all_cases.extend(cases)
            if cases and cases[0].source_sheet:
                sheets_used.append(cases[0].source_sheet)
        except Exception as exc:  # noqa: BLE001
            parse_errors.append(f"{path.name} [{sheet or 'default'}]: {exc}")

    jira_requirements = extract_requirements(jira_cache if jira_cache.exists() else None)
    jira_data = load_json(jira_cache) if jira_cache.exists() else {}
    confluence_payload = load_confluence_cache(issue_key)
    if not confluence_payload.get("ladrRequirements") and jira_data:
        try:
            confluence_payload = fetch_and_cache_confluence_for_issue(issue_key, jira_data, site=args.site)
        except RuntimeError:
            confluence_payload = load_confluence_cache(issue_key)
    ladr_requirements = dedupe_ladr_requirements(confluence_payload.get("ladrRequirements") or [])
    requirements = merge_requirement_sets(jira_requirements, ladr_requirements)

    for tc in all_cases:
        tc.steps = normalize_steps(tc.steps)
    if requirements and all_cases:
        map_testcases_to_requirements(all_cases, requirements)
    for tc in all_cases:
        tc.evidence_ids = extract_testcase_evidence_ids(tc, jira_requirements)

    coverage = compute_testplan_coverage(
        all_cases,
        requirements,
        jira_requirements=jira_requirements,
        ladr_requirements=ladr_requirements,
    )

    if all_cases:
        status = "ok"
    elif testplan_refs and not any(m.get("localFound") for m in attachment_meta if isinstance(m, dict)):
        status = "referenced_not_local"
    elif not files and not testplan_refs:
        status = "no_testplan"
    elif not files:
        status = "referenced_not_local"
    else:
        status = "parse_failed"

    primary_ref = testplan_refs[0] if testplan_refs else None
    if status == "referenced_not_local":
        coverage["testplanCoveragePct"] = "Pending"
        coverage["uncoveredRequirements"] = []
    filename = (primary_ref or {}).get("filename") or "Domino Test Plan.xlsx"
    sheet = (primary_ref or {}).get("sheet") or "Inc as full"
    excel_sheet = sheets_used[0] if sheets_used else sheet
    jira_att = next((m for m in attachment_meta if m.get("source") == "jira_attachment"), None)
    source_hint = "Jira attachment" if jira_att else ("local file" if files else "")
    coverage["coverageDetail"] = format_testplan_coverage_detail(coverage, source_hint)
    ladr_traceability = build_ladr_traceability(all_cases, ladr_requirements) if ladr_requirements else []
    scenario_filter = None
    if issue_key and all_cases:
        hay = " ".join(
            scenario_label(tc) + " " + " ".join(tc.steps.values()) for tc in all_cases
        ).lower()
        if "passport" in hay and "incremental" in hay:
            scenario_filter = "passport/incremental-as-full"
    summary_note = (
        format_testplan_summary_note(
            issue_key,
            filename,
            sheet,
            excel_sheet,
            len(all_cases),
            scenario_filter,
        )
        if status == "ok" and all_cases
        else None
    )
    payload = {
        "issueKey": issue_key,
        "fetchedAt": datetime.now(timezone.utc).isoformat(),
        "attachments": attachment_meta,
        "testPlanReferences": testplan_refs,
        "primaryReference": primary_ref,
        "filesParsed": [{"file": p.name, "sheet": s} for p, s in files],
        "sheetsUsed": sheets_used,
        "requirements": requirements,
        "jiraRequirements": jira_requirements,
        "ladrRequirements": ladr_requirements,
        "ladrTraceability": ladr_traceability,
        "confluence": {
            "status": confluence_payload.get("status"),
            "pages": [
                {
                    "title": p.get("title"),
                    "webUrl": p.get("webUrl"),
                    "pageId": p.get("id") or p.get("pageId"),
                    "error": p.get("error"),
                }
                for p in (confluence_payload.get("pages") or [])
            ],
        },
        "testCases": [asdict(tc) for tc in all_cases],
        "coverage": coverage,
        "testPlanSummaryNote": summary_note,
        "errors": parse_errors,
        "authError": auth_error,
        "status": status,
        "localSetupHint": (
            (
                f'Option C: attach "{filename}" to Jira and set ATLASSIAN_EMAIL + ATLASSIAN_API_TOKEN in .env '
                f'(see .env.example), then re-run fetch. '
                f'Or place file at testplans/{filename} (sheet: "{sheet}").'
            )
            if status == "referenced_not_local" and primary_ref
            else (
                f"Jira attachment download failed: {auth_error}. {credentials_hint()}"
                if auth_error and not all_cases
                else None
            )
        ),
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(out_path.resolve()),
                "status": status,
                "testCaseCount": len(all_cases),
                "references": len(testplan_refs),
                "primarySheet": (primary_ref.get("sheet") if primary_ref else None)
                or sheet
                if status == "referenced_not_local"
                else (primary_ref.get("sheet") if primary_ref else None),
            }
        )
    )
    return 0 if all_cases or status in {"referenced_not_local", "no_testplan"} else 1


if __name__ == "__main__":
    raise SystemExit(main())
